#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script d'import des panélistes Excel vers MySQL
Usage: python import_panel.py [fichier.xlsx]
"""

import sys
import json
import mysql.connector
from datetime import datetime, timedelta
from openpyxl import load_workbook

# Configuration base de données (à adapter selon votre environnement)
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'u486755141_etudes',
    'charset': 'utf8mb4'
}

# Durée avant expiration (1 mois)
EXPIRATION_DAYS = 30


def parse_date(jour, mois, annee):
    """Convertit JJ/MM/AA en date"""
    try:
        if jour and mois and annee:
            jour = int(jour)
            mois = int(mois)
            annee = int(annee)
            # Correction année sur 2 chiffres
            if annee < 100:
                annee = 1900 + annee if annee > 25 else 2000 + annee
            return f"{annee:04d}-{mois:02d}-{jour:02d}"
    except:
        pass
    return None


def parse_bool(value):
    """Convertit Oui/Non en boolean"""
    if value is None:
        return None
    val = str(value).strip().lower()
    return 1 if val in ['oui', 'yes', '1', 'true'] else 0


def clean_phone(phone):
    """Nettoie les numéros de téléphone"""
    if phone is None:
        return None
    phone = str(phone).strip()
    if phone in ['', 'None', '0']:
        return None
    # Ajoute le 0 devant si nécessaire
    if len(phone) == 9 and phone.isdigit():
        phone = '0' + phone
    return phone


def extract_children_data(row, headers):
    """Extrait les données des enfants"""
    children = []
    for i in range(1, 7):
        sexe_col = f'SEXE_ENFANT{i}'
        if sexe_col in headers:
            idx = headers.index(sexe_col)
            sexe = row[idx] if idx < len(row) else None
            if sexe:
                child = {
                    'sexe': sexe,
                    'date_naissance': parse_date(
                        row[headers.index(f'JJ_ENFANT{i}')] if f'JJ_ENFANT{i}' in headers else None,
                        row[headers.index(f'MM_ENFANT{i}')] if f'MM_ENFANT{i}' in headers else None,
                        row[headers.index(f'AA_ENFANT{i}')] if f'AA_ENFANT{i}' in headers else None
                    ),
                    'nom': row[headers.index(f'NOM_ENFANT{i}')] if f'NOM_ENFANT{i}' in headers else None,
                    'prenom': row[headers.index(f'PRENOM_ENFANT{i}')] if f'PRENOM_ENFANT{i}' in headers else None,
                    'accepte_test': parse_bool(row[headers.index(f'ACCEPT_IL_DE_PARTICIPER_AU_TEST_ENFANT{i}')] if f'ACCEPT_IL_DE_PARTICIPER_AU_TEST_ENFANT{i}' in headers else None)
                }
                children.append(child)
    return children if children else None


def extract_conjoint_data(row, headers):
    """Extrait les données du conjoint"""
    civilite = row[headers.index('CIVILITE_CONJOINT')] if 'CIVILITE_CONJOINT' in headers else None
    if not civilite:
        return None

    return {
        'civilite': civilite,
        'nom': row[headers.index('NOM_CONJOINT')] if 'NOM_CONJOINT' in headers else None,
        'prenom': row[headers.index('PRENOM_CONJOINT')] if 'PRENOM_CONJOINT' in headers else None,
        'tel_domicile': clean_phone(row[headers.index('TEL_DOMICILE_CONJOINT')] if 'TEL_DOMICILE_CONJOINT' in headers else None),
        'tel_portable': clean_phone(row[headers.index('TEL_PORTABLE_CONJOINT')] if 'TEL_PORTABLE_CONJOINT' in headers else None),
        'tel_bureau': clean_phone(row[headers.index('TEL_BUREAU_CONJOINT')] if 'TEL_BUREAU_CONJOINT' in headers else None),
        'email': row[headers.index('EMAIL_CONJOINT')] if 'EMAIL_CONJOINT' in headers else None,
        'date_naissance': parse_date(
            row[headers.index('JJ_CONJOINT')] if 'JJ_CONJOINT' in headers else None,
            row[headers.index('MM_CONJOINT')] if 'MM_CONJOINT' in headers else None,
            row[headers.index('AA_CONJOINT')] if 'AA_CONJOINT' in headers else None
        ),
        'diplome': row[headers.index('DIPLOME_OBTENU_CONJOINT')] if 'DIPLOME_OBTENU_CONJOINT' in headers else None,
        'situation_professionnelle': row[headers.index('SITUATION_PROFESSIONNEL_CONJOINT')] if 'SITUATION_PROFESSIONNEL_CONJOINT' in headers else None,
        'profession': row[headers.index('PROFESSION_ACTUEl_CONJOINT')] if 'PROFESSION_ACTUEl_CONJOINT' in headers else None,
        'accepte_test': parse_bool(row[headers.index('ACCEPT_IL_DE_PARTICIPER_AU_TEST')] if 'ACCEPT_IL_DE_PARTICIPER_AU_TEST' in headers else None),
        'banque_principale': row[headers.index('BANQUE_PRINCIPAL_CONJOINT')] if 'BANQUE_PRINCIPAL_CONJOINT' in headers else None,
        'autres_banques': row[headers.index('AUTRES_BANQUES_CONJOINT')] if 'AUTRES_BANQUES_CONJOINT' in headers else None,
    }


def extract_vehicles_data(row, headers, prefix='VOITURE'):
    """Extrait les données des véhicules"""
    vehicles = []
    for i in range(1, 4):
        modele_col = f'MODELE_{prefix}' if i == 1 else f'MODELE_{prefix}{i}'
        if modele_col in headers:
            idx = headers.index(modele_col)
            modele = row[idx] if idx < len(row) else None
            if modele and str(modele).strip():
                vehicle = {
                    'modele': modele,
                    'annee_achat': row[headers.index(f'ANNEE_ACHAT_{prefix}{i}')] if f'ANNEE_ACHAT_{prefix}{i}' in headers else None,
                    'type': row[headers.index(f'TYPE_DE_VEHICULE_{prefix}{i}')] if f'TYPE_DE_VEHICULE_{prefix}{i}' in headers else None,
                    'type_achat': row[headers.index(f'TYPE_ACHAT_{prefix}{i}')] if f'TYPE_ACHAT_{prefix}{i}' in headers else None,
                }
                vehicles.append(vehicle)
    return vehicles if vehicles else None


def extract_equipements(row, headers):
    """Extrait les équipements"""
    equipements = {}
    mapping = {
        'POSSEDE_TELEVISEUR': 'televiseur',
        'POSSEDE_VIDEO_PROJECTEUR': 'video_projecteur',
        'POSSEDE_UN_ORDINATEUR_PORTABLE': 'ordinateur_portable',
        'POSSEDE_APPAREIL_PHOTO_NUMERIQUE': 'appareil_photo',
        'POSSEDE_WIFI': 'wifi',
        'POSSEDE_GPS': 'gps'
    }
    for col, key in mapping.items():
        if col in headers:
            val = row[headers.index(col)]
            if val:
                equipements[key] = parse_bool(val)
    return equipements if equipements else None


def import_excel_to_mysql(excel_file):
    """Importe le fichier Excel dans MySQL"""
    print(f"Chargement de {excel_file}...")
    wb = load_workbook(excel_file, read_only=True)
    ws = wb.active

    # Récupère les en-têtes
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else f'COL_{len(headers)}')

    print(f"Colonnes: {len(headers)}")
    print(f"Lignes: {ws.max_row - 1}")

    # Connexion MySQL
    print("Connexion à MySQL...")
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # Date d'expiration (1 mois après aujourd'hui)
    expires_at = (datetime.now() + timedelta(days=EXPIRATION_DAYS)).strftime('%Y-%m-%d %H:%M:%S')

    # Préparation de l'insertion
    insert_sql = """
        INSERT INTO panel_imported (
            panel_id, email, region, civilite, nom, prenom,
            adresse, code_postal, departement, ville,
            tel_domicile, tel_portable, tel_bureau,
            date_naissance, age, situation_familiale, situation_professionnelle,
            enfants_au_foyer, enfants_plus_18ans, enfants_moins_18ans, enfants_data,
            diplome, profession, secteur_activite, revenu_mensuel,
            banque_principale, autres_banques, conjoint_data,
            type_habitation, situation_habitation,
            possede_voiture, voitures_data, possede_moto, motos_data,
            equipements, expires_at
        ) VALUES (
            %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s,
            %s, %s,
            %s, %s, %s, %s,
            %s, %s
        ) ON DUPLICATE KEY UPDATE
            region = VALUES(region),
            civilite = VALUES(civilite),
            nom = VALUES(nom),
            prenom = VALUES(prenom),
            expires_at = VALUES(expires_at)
    """

    imported = 0
    skipped = 0
    errors = 0

    print("Import en cours...")

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Vérifie les champs obligatoires
            panel_id = row[headers.index('ID')] if 'ID' in headers else None
            email = row[headers.index('EMAIL')] if 'EMAIL' in headers else None

            if not panel_id or not email:
                skipped += 1
                continue

            # Nettoyage email
            email = str(email).strip().lower()
            if '@' not in email:
                skipped += 1
                continue

            # Extraction des données
            data = (
                int(panel_id),
                email,
                row[headers.index('REGION')] if 'REGION' in headers else None,
                row[headers.index('CIVILITE')] if 'CIVILITE' in headers else None,
                row[headers.index('NOM')] if 'NOM' in headers else None,
                row[headers.index('PRENOM')] if 'PRENOM' in headers else None,
                row[headers.index('ADRESSE_POSTAL')] if 'ADRESSE_POSTAL' in headers else None,
                str(row[headers.index('CODE_POSTAL')]) if 'CODE_POSTAL' in headers and row[headers.index('CODE_POSTAL')] else None,
                str(row[headers.index('DEPARTEMENT')]) if 'DEPARTEMENT' in headers and row[headers.index('DEPARTEMENT')] else None,
                row[headers.index('VILLE')] if 'VILLE' in headers else None,
                clean_phone(row[headers.index('TEL_DOMICILE')] if 'TEL_DOMICILE' in headers else None),
                clean_phone(row[headers.index('TEL_PORTABLE')] if 'TEL_PORTABLE' in headers else None),
                clean_phone(row[headers.index('TEL_BUREAU')] if 'TEL_BUREAU' in headers else None),
                parse_date(
                    row[headers.index('JJ')] if 'JJ' in headers else None,
                    row[headers.index('MM')] if 'MM' in headers else None,
                    row[headers.index('AA')] if 'AA' in headers else None
                ),
                int(row[headers.index('AGE')]) if 'AGE' in headers and row[headers.index('AGE')] else None,
                row[headers.index('SITUATION_FAMILIAL')] if 'SITUATION_FAMILIAL' in headers else None,
                row[headers.index('SITUATION_PROFESSIONNEL')] if 'SITUATION_PROFESSIONNEL' in headers else None,
                int(row[headers.index('ENFANTS_AU_FOYER')] or 0) if 'ENFANTS_AU_FOYER' in headers else 0,
                int(row[headers.index('ENFANTS_PLUS_DE_18ANS')] or 0) if 'ENFANTS_PLUS_DE_18ANS' in headers else 0,
                int(row[headers.index('ENFANTS_MOINS_DE_18ANS')] or 0) if 'ENFANTS_MOINS_DE_18ANS' in headers else 0,
                json.dumps(extract_children_data(row, headers), ensure_ascii=False) if extract_children_data(row, headers) else None,
                row[headers.index('DIPLOME_OBTENU')] if 'DIPLOME_OBTENU' in headers else None,
                row[headers.index('PROFESSION_ACTUEL')] if 'PROFESSION_ACTUEL' in headers else None,
                row[headers.index('SECTEUR_ACTIVITE')] if 'SECTEUR_ACTIVITE' in headers else None,
                row[headers.index('REVENU_NET_MENSUEL_DU_FOYER')] if 'REVENU_NET_MENSUEL_DU_FOYER' in headers else None,
                row[headers.index('BANQUE_PRINCIPAL')] if 'BANQUE_PRINCIPAL' in headers else None,
                row[headers.index('AUTRES_BANQUES')] if 'AUTRES_BANQUES' in headers else None,
                json.dumps(extract_conjoint_data(row, headers), ensure_ascii=False) if extract_conjoint_data(row, headers) else None,
                row[headers.index('TYPE_HABITATION')] if 'TYPE_HABITATION' in headers else None,
                row[headers.index('SITUATION_HABITATION')] if 'SITUATION_HABITATION' in headers else None,
                parse_bool(row[headers.index('VOITURE')] if 'VOITURE' in headers else None),
                json.dumps(extract_vehicles_data(row, headers, 'VOITURE'), ensure_ascii=False) if extract_vehicles_data(row, headers, 'VOITURE') else None,
                parse_bool(row[headers.index('MOTO')] if 'MOTO' in headers else None),
                json.dumps(extract_vehicles_data(row, headers, 'MOTO'), ensure_ascii=False) if extract_vehicles_data(row, headers, 'MOTO') else None,
                json.dumps(extract_equipements(row, headers), ensure_ascii=False) if extract_equipements(row, headers) else None,
                expires_at
            )

            cursor.execute(insert_sql, data)
            imported += 1

            if imported % 1000 == 0:
                conn.commit()
                print(f"  {imported} enregistrements importés...")

        except Exception as e:
            errors += 1
            if errors <= 10:
                print(f"  Erreur ligne {row_num}: {e}")

    conn.commit()
    cursor.close()
    conn.close()

    print("\n=== RÉSUMÉ ===")
    print(f"Importés: {imported}")
    print(f"Ignorés (sans email/ID): {skipped}")
    print(f"Erreurs: {errors}")
    print(f"Expiration: {expires_at}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        excel_file = 'panel_clean.xlsx'
    else:
        excel_file = sys.argv[1]

    import_excel_to_mysql(excel_file)
